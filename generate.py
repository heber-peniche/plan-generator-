#!/usr/bin/env python3
"""
Generador de planes de trabajo — Volumetrics by AIVARA.

Toma los 6 datos por cliente (contribuyente, instalaciones, molécula,
partidas a excluir, unidad verificadora, fecha de kickoff) e inyecta un
plan de trabajo HTML autocontenido (sin dependencias externas de terceros,
salvo la fuente Inter vía Google Fonts, con fallback a system-ui offline).

Uso interactivo:
    python generate.py

Uso no interactivo:
    python generate.py --contribuyente "General Motors" --instalaciones 4 \
        --molecula "Gas Natural" --excluir Certificación \
        --unidad-verificadora MG3 --kickoff 2026-09-07

Las partidas fijas (A-E, con sus subtareas y duraciones) viven en
data/partidas.json y no cambian entre clientes — están tomadas del Excel
"Template - Plan de trabajo y Matriz Responsabilidad - Controles Volumetricos".

Opcionalmente, --matriz-excel importa la hoja "Matriz de Responsabilidades" de
ese mismo Excel (RACI por tarea) y agrega una pestaña extra al plan generado.
Requiere `pip install openpyxl` (no es una dependencia del resto del script).
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "templates" / "plan_template.html"
PARTIDAS_PATH = BASE_DIR / "data" / "partidas.json"
OUTPUT_DIR = BASE_DIR.parent / "output"

# id de partida -> nombres/alias reconocidos para "partidas a excluir"
PARTIDA_ALIAS = {
    "A": ["a", "pre-auditoria", "pre auditoria", "preauditoria", "diagnostico"],
    "B": ["b", "reportes historicos", "historicos", "sat 2022-2025", "reportes sat"],
    "C": ["c", "programa informatico", "software", "nube", "covol"],
    "D": ["d", "sgm", "sgm 10012", "balance volumetrico"],
    "E": ["e", "certificacion", "certificado", "certificado anual"],
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def resolver_partidas_excluir(valores):
    """Acepta ids ('E'), nombres ('Certificación') o listas mixtas; case/accent-insensitive."""
    ids = set()
    desconocidos = []
    for v in valores:
        v = v.strip()
        if not v:
            continue
        nv = _norm(v)
        encontrado = None
        for pid, alias in PARTIDA_ALIAS.items():
            # Solo se hace matching por substring en alias de 3+ caracteres,
            # para que letras sueltas ("a", "e") no den falsos positivos por
            # ser substring de un nombre largo (ej. "certificación" contiene "a").
            if nv == pid.lower() or any(
                nv == a or (len(a) >= 3 and (nv in a or a in nv)) for a in alias
            ):
                encontrado = pid
                break
        if encontrado:
            ids.add(encontrado)
        else:
            desconocidos.append(v)
    if desconocidos:
        print(f"Aviso: no reconozco estas partidas a excluir, se ignoran: {', '.join(desconocidos)}", file=sys.stderr)
    return sorted(ids)


def resolver_instalaciones(valor: str):
    """Si es un número puro -> 'Instalación 1..N'; si no, se toma como lista separada por comas."""
    valor = (valor or "").strip()
    if valor.isdigit():
        n = int(valor)
        return [f"Instalación {i}" for i in range(1, n + 1)]
    nombres = [s.strip() for s in valor.split(",") if s.strip()]
    return nombres or ["Instalación 1"]


def slugify(s: str) -> str:
    s = _norm(s)
    return re.sub(r"\s+", "-", s).strip("-") or "cliente"


LETRAS_RACI = {"R", "A", "C", "I"}


def parse_matriz_excel(path: Path) -> dict:
    """
    Lee la hoja "Matriz de Responsabilidades" del Excel de plan de trabajo
    (mismo layout que "Template - Plan de trabajo y Matriz Responsabilidad -
    Controles Volumetricos.xlsx"): fila de encabezado con "TASK" en A, columnas
    de roles a partir de C, filas de tarea con letras R/A/C/I, y al final una
    leyenda con una letra por fila en la columna B.

    No asume que las letras estén siempre en la fila de la partida "padre":
    se toma literalmente cualquier fila con texto en la columna A que tenga
    al menos una letra RACI asignada, tal como esté capturado en el Excel.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "Para usar --matriz-excel instala openpyxl primero: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(path, data_only=True)
    nombre_hoja = next((s for s in wb.sheetnames if "matriz" in _norm(s)), None)
    if not nombre_hoja:
        raise ValueError(
            f"No encontré una hoja de 'Matriz de responsabilidad' en {path.name} "
            f"(hojas disponibles: {', '.join(wb.sheetnames)})"
        )
    ws = wb[nombre_hoja]

    fila_encabezado = None
    for row in ws.iter_rows(min_row=1, max_row=15):
        if row[0].value and str(row[0].value).strip().upper() == "TASK":
            fila_encabezado = row[0].row
            break
    if fila_encabezado is None:
        raise ValueError(f"No encontré la fila de encabezado ('TASK' en la columna A) en la hoja '{nombre_hoja}'.")

    columnas, columnas_idx = [], []
    for col in range(3, ws.max_column + 1):
        val = ws.cell(row=fila_encabezado, column=col).value
        texto = str(val).strip() if val not in (None, "") else ""
        if texto:
            columnas.append(texto)
            columnas_idx.append(col)

    filas, leyenda = [], []
    for r in range(fila_encabezado + 1, ws.max_row + 1):
        tarea = ws.cell(row=r, column=1).value
        asignado = ws.cell(row=r, column=2).value
        valores = [ws.cell(row=r, column=c).value for c in columnas_idx]
        valores_txt = [str(v).strip() if v not in (None, "") else "" for v in valores]

        if tarea and any(valores_txt):
            filas.append({
                "tarea": str(tarea).strip(),
                "responsable": str(asignado).strip() if asignado not in (None, "") else "",
                "valores": valores_txt,
            })
        elif not tarea and asignado and str(asignado).strip().upper() in LETRAS_RACI:
            titulo = ws.cell(row=r, column=3).value
            descripcion = ws.cell(row=r, column=4).value
            if titulo:
                leyenda.append({
                    "letra": str(asignado).strip().upper(),
                    "titulo": str(titulo).strip(),
                    "descripcion": str(descripcion).strip() if descripcion not in (None, "") else "",
                })

    if not filas:
        raise ValueError(f"La hoja '{nombre_hoja}' no tiene filas con letras RACI capturadas.")

    return {"columnas": columnas, "filas": filas, "leyenda": leyenda}


def construir_client_data(contribuyente, instalaciones_raw, molecula, excluir_raw,
                           unidad_verificadora, kickoff, matriz=None):
    instalaciones = resolver_instalaciones(instalaciones_raw)
    excluir_ids = resolver_partidas_excluir(
        excluir_raw if isinstance(excluir_raw, list) else re.split(r"[,;]", excluir_raw or "")
    )
    overrides = {}
    if excluir_ids:
        for i in range(len(instalaciones)):
            overrides[str(i)] = {"excluir": excluir_ids}

    client_data = {
        "contribuyente": contribuyente.strip(),
        "unidadVerificadora": unidad_verificadora.strip(),
        "moleculaGeneral": molecula.strip(),
        "fechaKickoff": kickoff.strip(),
        "instalaciones": ", ".join(instalaciones),
        "overrides": overrides,
        # Aplica también a instalaciones agregadas en vivo desde el HTML (botón
        # "+ Instalación"), para que hereden la misma exclusión que las demás.
        "excluirGeneral": excluir_ids,
    }
    if matriz:
        client_data["matriz"] = matriz
    return client_data


def render(client_data: dict) -> str:
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    partidas_data = json.loads(PARTIDAS_PATH.read_text(encoding="utf-8"))
    html = template.replace(
        "{{ client_json }}", json.dumps(client_data, ensure_ascii=False, indent=2)
    ).replace(
        "{{ partidas_json }}", json.dumps(partidas_data, ensure_ascii=False)
    ).replace(
        "{{ contribuyente }}", client_data["contribuyente"]
    )
    return html


def generar(contribuyente, instalaciones, molecula, excluir, unidad_verificadora, kickoff,
            out_dir: Path = OUTPUT_DIR, matriz_excel: Path = None) -> Path:
    matriz = parse_matriz_excel(matriz_excel) if matriz_excel else None
    client_data = construir_client_data(
        contribuyente, instalaciones, molecula, excluir, unidad_verificadora, kickoff, matriz
    )
    html = render(client_data)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{slugify(contribuyente)}-plan-de-trabajo.html"
    out_path.write_text(html, encoding="utf-8")
    return out_path


def parse_args():
    p = argparse.ArgumentParser(description="Genera un plan de trabajo HTML por cliente.")
    p.add_argument("--contribuyente", help="Nombre de la empresa, ej. 'General Motors'")
    p.add_argument("--instalaciones", help="Número de instalaciones (ej. 4) o nombres separados por comas")
    p.add_argument("--molecula", help="Molécula / producto, ej. 'Gas Natural'")
    p.add_argument("--excluir", default="", help="Partidas a excluir (ids o nombres), separadas por comas, ej. 'Certificación'")
    p.add_argument("--unidad-verificadora", dest="unidad_verificadora", help="Ej. MG3")
    p.add_argument("--kickoff", help="Fecha de kickoff, ej. '2026-09-07' o '2026-09-07, Planta 2: 2026-10-01'")
    p.add_argument("--out-dir", default=str(OUTPUT_DIR), help="Carpeta de salida")
    p.add_argument("--matriz-excel", help="Ruta al Excel con la hoja 'Matriz de Responsabilidades' (opcional, requiere openpyxl)")
    return p.parse_args()


def prompt(msg, default=None):
    suffix = f" [{default}]" if default else ""
    val = input(f"{msg}{suffix}: ").strip()
    return val or (default or "")


def main():
    args = parse_args()
    if not args.contribuyente:
        print("=== Generador de planes de trabajo — Volumetrics by AIVARA ===\n")
        contribuyente = prompt("Contribuyente (nombre de la empresa)")
        instalaciones = prompt("Número de instalaciones (o nombres separados por comas)", "1")
        molecula = prompt("Molécula / producto", "Gas L.P.")
        excluir = prompt("Partidas a excluir (separadas por comas, Enter si ninguna)", "")
        unidad_verificadora = prompt("Unidad verificadora del proyecto", "VICER")
        kickoff = prompt("Fecha de kickoff (YYYY-MM-DD)")
        matriz_excel = prompt("Excel con la Matriz de Responsabilidad (Enter para omitir)", "")
        excluir_list = [s.strip() for s in excluir.split(",") if s.strip()]
    else:
        contribuyente = args.contribuyente
        instalaciones = args.instalaciones or "1"
        molecula = args.molecula or "Gas L.P."
        unidad_verificadora = args.unidad_verificadora or "VICER"
        kickoff = args.kickoff or ""
        matriz_excel = args.matriz_excel or ""
        excluir_list = args.excluir

    if not contribuyente:
        print("Error: el contribuyente es obligatorio.", file=sys.stderr)
        sys.exit(1)
    if not kickoff:
        print("Error: la fecha de kickoff es obligatoria (YYYY-MM-DD).", file=sys.stderr)
        sys.exit(1)

    out_path = generar(
        contribuyente, instalaciones, molecula, excluir_list, unidad_verificadora, kickoff,
        out_dir=Path(args.out_dir) if hasattr(args, "out_dir") and args.out_dir else OUTPUT_DIR,
        matriz_excel=Path(matriz_excel) if matriz_excel else None,
    )
    print(f"\nListo: {out_path}")


if __name__ == "__main__":
    main()
